from django.http import JsonResponse
from django.views import View

from movies.models import Movie


class CreateLike(View):
    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if 'movie_id' in kwargs['movie_id']:
                parent_user = request.user
                sub_user = parent_user.sub_user.all().filter(logined=True).get()
                movie = Movie.objects.get(pk=kwargs['movie_id'])
                if sub_user in movie.likes.all():
                    movie.likes.remove(sub_user)
                    return JsonResponse({'data': 'remove'})
                else:
                    movie.likes.add(sub_user)
                    return JsonResponse({'data': 'add'})


def update_real(request):
    movies = Movie.objects.all()
    for movie in movies:
        runningtime = movie.running_time
        if '시간 ' in runningtime:
            runningtime = runningtime.split('시간 ')
            total_minute = int(runningtime[0]) * 60 + int(runningtime[1][:-1])
        else:
            total_minute = int(runningtime[:-1])
        movie.real_running_time = total_minute * 60
        movie.save()


from django.contrib.auth import get_user_model
from django.http import JsonResponse, HttpResponse
from django.views import View
from rest_framework import generics
from rest_framework.response import Response

from accounts.models import SubUser, LikeDisLikeMarked
from .serializers import *

import time
import re

from django.shortcuts import render
from selenium.webdriver.chrome.options import Options

from bs4 import BeautifulSoup
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from openpyxl import Workbook
from .models import *

from openpyxl import load_workbook


def save_data1(request):
    def save_db(model, value_list, movie_instance):
        if len(value_list) > 1:
            for value in value_list:
                model_instance = model.objects.get_or_create(name=value)[0]
                model_instance.movie.add(movie_instance)

        elif not value_list:
            return

        elif len(value_list) == 1:
            model_instance = model.objects.get_or_create(name=value_list[0])[0]
            model_instance.movie.add(movie_instance)

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_5)\
                AppleWebKit 537.36 (KHTML, like Gecko) Chrome")

    driver = webdriver.Chrome('/home/park/Downloads/chromedriver', options=options)

    driver.implicitly_wait(3)

    # 브라우저 오픈
    driver.get('http://www.netflix.com/browse')
    time.sleep(10)

    # 로그인
    driver.find_element_by_name('userLoginId').send_keys('roqkfwkehlwk@naver.com')
    time.sleep(3)
    driver.find_element_by_name('password').send_keys('admin67890!@')
    time.sleep(3)
    driver.find_element_by_xpath('//*[@class="btn login-button btn-submit btn-small"]').click()
    time.sleep(3)
    # 서브계정 선택
    driver.find_element_by_xpath(
        '//*[@style="background-image:url(https://occ-0-2794-2219.1.nflxso.net/art/0d6b9/6c139bd2c216b7608148431203765747ce20d6b9.png)"]').click()
    time.sleep(10)

    wb2 = load_workbook('fast_flix2.xlsx')
    ws = wb2.active

    for row in ws.values:
        if row[0]:
            pk = row[0]
            mov = Movie.objects.get(pk=pk)
            category = Genre.objects.get_or_create(name=row[3])[0]
            mov.genre.add(category)
            mov.save()

        else:
            driver.find_element_by_xpath('//*[@class="icon-search"]').click()
            time.sleep(10)
            driver.find_element_by_xpath('//input[@data-uia="search-box-input"]').send_keys(row[1])
            time.sleep(10)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            page1 = soup.select('div.title-card-container .ptrack-content')

            title = page1[0].get_text()
            if Movie.objects.filter(name=title):
                print(f'{title}--------- 생략')
                driver.find_element_by_xpath('//*[@class="icon-close"]').click()
                time.sleep(10)
                driver.find_element_by_xpath('//*[@class="icon-search"]').click()
                time.sleep(10)
                continue
            else:
                print('진행')

            # 가로그림
            horizontal_image = page1[0].find('img')['src']

            page_a = page1[0].find_all('a')
            href = page_a[0]['href']

            movie_num = int(href.split('?')[0].split('/watch/')[1])
            driver.get(f'https://www.netflix.com/title/{movie_num}')
            time.sleep(10)

            # 시놉시스
            synopsis = driver.find_element_by_xpath('//div[@class="synopsis"]').text
            # 러닝타임
            running_time = driver.find_element_by_xpath('//*[@class="duration"]').text
            # 개봉년도
            pro_year = driver.find_element_by_xpath('//span[@class="year"]').text
            # 큰 사이즈 그림
            big_image_src = driver.find_element_by_xpath('//div[@class="image-rotator-image "]').get_attribute('style')
            p = re.compile('http.*"')

            big_image = p.findall(big_image_src)[0][:-1]

            driver.find_element_by_xpath('//*[@class="ShowDetails"]').click()
            time.sleep(10)
            # 관람가
            maturity = driver.find_element_by_xpath('//span[@class="maturityDescription"]').text

            time.sleep(10)
            # 로고
            logo = driver.find_element_by_xpath('//*[@class="logo small-logo"]').get_attribute('src')
            time.sleep(10)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            page22 = soup.select('div.detailsItem.detailsMedia')

            page2 = ''
            if len(page22) > 1:
                for page_num in range(len(page22)):
                    page2 += '/' + page22[page_num].get_text("/")
            else:
                page2 = page22[0].get_text("/")

            # 감독
            directors = []
            # 배우
            actors = []
            # 각본
            authors = []
            # 제작
            company = []

            change = []

            for j in page2.split('/'):
                j = j.rstrip()
                if j == '제작':
                    change = company
                    continue
                if j == '감독':
                    change = directors
                    continue
                if j == '출연':
                    change = actors
                    continue
                if j == '각본':
                    change = authors
                    continue
                change.append(j)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            page3 = soup.select('div.detailsItem.detailsTags')[0].get_text("/")

            # 장르
            genres = []
            # 특징
            features = []
            change = genres

            for k in page3.split('/')[1:]:
                k = k.rstrip()
                if k == '영화 특징':
                    change = features
                    continue
                change.append(k)

            vertical_image = row[2]

            special_category = Genre.objects.get_or_create(name=row[3])[0]
            genres.append(special_category)

            data_list = [title, synopsis, running_time, pro_year, maturity, logo, directors, actors, authors, genres,
                         features, big_image]

            normal = [title, synopsis, running_time, pro_year, logo, features]
            refer = [directors, actors, authors, genres]
            # maturity는 foreign_key

            for i in normal:
                print(i)

            for j in refer:
                print(j)

            movie_instance = Movie.objects.create(name=title, synopsis=synopsis, running_time=running_time,
                                                  production_date=pro_year, logo_image_path=logo,
                                                  big_image_path=big_image, horizontal_image_path=horizontal_image,
                                                  vertical_image=vertical_image)

            save_db(Director, directors, movie_instance)
            save_db(Actor, actors, movie_instance)
            save_db(Author, authors, movie_instance)
            save_db(Genre, genres, movie_instance)
            save_db(Feature, features, movie_instance)

            degree = Degree.objects.get_or_create(name=maturity)[0]
            movie_instance.degree = degree
            movie_instance.save()

            # 영화버튼 클릭
            # driver.find_element_by_xpath('//*[@href="/browse/genre/34399"]').click()

            # 엑셀 열구분

            # for i in range(20):
            #
            #     column_count = 1

            # 제목

            # for data in data_list:
            #     if type(data) == list:
            #         data = ','.join(data)
            #     ws.cell(row=i + 5, column=column_count, value=data)
            #     column_count += 1
            #     wb.save('movies1.xlsx')

    return HttpResponse({})
